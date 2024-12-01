
import logging
import json
import mimetypes
import shutil
import uuid
import io
import os

from fastapi import APIRouter, Request
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from PIL import Image as PilImage

from app.conf import const
from app.util import util_library, util_panel, util_response, util_param, util_db, util_file

logger = logging.getLogger()
router = APIRouter()

@router.post("/excel")
async def upload_file(request: Request):

    form = await request.form()
    file = form.getlist('file')[0]
 
    request_param = util_param.get_init_info ( request, only_params=True )
    file_ext = util_file.get_extension(file.filename)
    if file_ext != "xlsx" : raise(Exception("invalid file: only support xlsx file"))

    file_path =  util_file.make_directory(f"{const.PATH_DATA_UPLOAD}/{request_param["@id"]}/")
    file_path += str(uuid.uuid4()).replace("-", "")
    file_path += ".xlsx"

    with open( file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        file.file.close()


    if "target-sub" in form and  form["target-sub"] == "bulk" :

        grp = form["g"]
        idx = form["i"]
        entity = form["entity"]
        mode = form["mode"]
        target = form["target"]
        run  = form["run"]

        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.sheetnames[0]
        worksheet = workbook[sheet]

        data = []
        columns = [cell.value for cell in worksheet[1]]
        for record in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {columns[col_idx]: cell for col_idx, cell in enumerate(record)}
            if any(value is not None for value in row_data.values()): data.append(row_data)
        util_file.delete_file(file_path)

        post = {
            "g": grp,
            "i": idx,
            "entity": entity,
            "mode": mode,
            "target": target,
            "run": run,
            "@data": {}
        }
        post["@data"]["new"] = data

        panel, panel_json, params = util_param.get_init_info(request, post)
        if panel == None : return util_response.error("invalid access")

        return await util_panel.execute_panel (panel_json, params, logger)

    else :
        custom = {}
        custom[".g"] = form["g"]
        custom[".i"] = form["i"]
        for key, val in form.items():
            if key not in ["g", "i", "file"] : custom[key] = val

        params = {**request_param, **custom}

        panel_db = util_panel.get_panel_db (params) 
        if panel_db is None : return util_response.error("invalid access")
        panel_json = json.loads(panel_db["json_panel_value"])

        info = util_library.get_obj_array(panel_json[params["entity"]][params["mode"]] , "name", params["target"])
        info["grp"] = params[".g"]
        info["idx"] = params[".i"]
        info["user"] = params["@id"]

        prework = None
        if "prework" in info:
            prework = {}
            prework["query"] = info["prework"]["query"]
            prework["params"] = [params]

        if info["type"] == "excel" :

            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.sheetnames[0]
            worksheet = workbook[sheet]

            data_pils = []
            if "image" in info :
                for image in worksheet._images:
                    image_data = io.BytesIO(image._data())
                    pil_image = PilImage.open(image_data)
                    data_pils.append(pil_image)

            data_org = []
            data_new = []
            cnt = 0
            columns = [cell.value for cell in worksheet[1]]
            for record in worksheet.iter_rows(min_row=2, values_only=True):
                row_data = {columns[col_idx].strip(): cell for col_idx, cell in enumerate(record)}
                data_org.append(row_data)

                row = {}
                for k, v in info["ecolumns"].items() :
                    if k in row_data : 
                        row[v] = row_data[k]
                    elif "defaults" in info and k in info["defaults"] : 
                        row[v] = info["defaults"][k]

                if "post" in info :
                    for k, v in info["post"].items() :
                        for code in v :
                            row[k] = eval(code)

                if "image" in info :
                    for image_info in info["image"] :
                        image_name = row[image_info["key"]] + ".png"
                        image_path = os.path.join(const.PATH_DATA_UPLOAD, params["@id"], image_name)
                        data_pils[cnt].save(image_path)

                        if (image_info["type"] == "gcs") :
                                
                            upload_info = {}
                            upload_info["bucket"] = image_info["bucket"]
                            upload_info["filename"] = image_name
                            upload_info["idx"] = ""
                            upload_info["path_cloud"] = util_db.get_parsed_query(image_info["path_cloud"], row)
                            upload_info["path_url"] = util_db.get_parsed_query(image_info["path_url"], row)

                            file_info = {}
                            file_info["path"] = image_path
                            file_info["name"] = row[image_info["key"]]
                            file_info["ext"] = "png"
                            file_info["date"] = ""
                            file_info["time"] = ""
                            file_info["mime"] = "image/png"
                            
                            upload_res = util_file.upload_gcs (upload_info, file_info)
                            row[image_info["name"]] = upload_res["url"]
                        

                data_new.append( {**params, **row} )
                cnt += 1
        
            if "query" in info : 
                util_db.execute_db (panel_json["datasource"], info["query"], data_new, prework=prework)
                util_file.delete_file(file_path)

        final_res = {}
        final_res["status"] = "ok"
        final_res["grp"] = params[".g"]
        final_res["pid"] = params[".i"]
        final_res["entity"] = params["entity"]
        final_res["mode"] = params["mode"]
        final_res["type"] = params["type"]
        final_res["target"] = params["target"]

        #util_library.log(logger, params, data_new)

        return final_res

@router.post("/upload")
async def upload_file(request: Request):

    custom = {}
    form = await request.form()

    custom[".g"] = form["g"]
    custom[".i"] = form["i"]
    custom["entity"] = form["entity"]
    custom["mode"] = form["mode"]
    custom["type"] = form["type"]
    custom["target"] = form["target"]
    files = form.getlist('files[]')

    request_param = util_param.get_init_info ( request, only_params=True )
    params = {**request_param, **custom}

    try :
        panel_db = util_panel.get_panel_db (params) 
        if panel_db is None : return util_response.error("invalid access")
        panel_json = json.loads(panel_db["json_panel_value"])

        info = panel_json[params["entity"]]["uploads"][params["target"]]
        info["grp"] = params[".g"]
        info["idx"] = params[".i"]
        info["user"] = params["@id"]
        info["upload_path"] = const.PATH_DATA_UPLOAD
        info["asset_path"] = const.PATH_DATA_ASSET

    except Exception:
        raise(Exception("invalid param 1"))
    
    res_arr = []
    for file in files :
        if "accept" in info :            
            file_ext = util_file.get_extension(file.filename)
            if f".{file_ext}" not in info["accept"] : raise(Exception("invalid param 2"))

        res_obj = util_file.upload(info, file)
        res_arr.append(res_obj)

    final_res = {}
    final_res["status"] = "ok"
    final_res["grp"] = params[".g"]
    final_res["pid"] = params[".i"]
    final_res["entity"] = params["entity"]
    final_res["mode"] = params["mode"]
    final_res["type"] = params["type"]
    final_res["target"] = params["target"]
    final_res["data"] = res_arr

    return final_res


@router.get("/download")
async def downlad_file(request: Request):

    params = util_param.get_init_info(request, only_params = True)
    if ".p" not in params : return util_response.error("invalid access")

    file_path = f"{const.PATH_DATA_ASSET}/{params['.p']}"
    file_name = file_path.split('/')[-1]

    mime_type, encoding = mimetypes.guess_type(file_name)

    return FileResponse(
        path=file_path,
        media_type=mime_type,
        headers={"Content-Disposition": f"attachment; filename={file_name}"}
    )

@router.get("/image")
async def image_file(request: Request):

    params = util_param.get_init_info(request, only_params = True)
    if ".p" not in params : return util_response.error("invalid access")

    file_path = f"{const.PATH_DATA_ASSET}/{params['.p']}"
    file_name = file_path.split('/')[-1]

    mime_type, encoding = mimetypes.guess_type(file_name)

    return FileResponse(
        path=file_path,
        media_type=mime_type
    )
